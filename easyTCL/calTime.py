# -*- coding: utf-8 -*-
"""
Created on Thu Jul  5 13:24:09 2018

@author: 09305
"""


import openpyxl

class tclRoom:
    def __init__(self):
       self.value = 0

    def calTimeB(self, number):
        self.number = number
        wb = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\data_B.xlsx', data_only=True)
        ws = wb['datasheet']
        rawList = []
        for i in range(1, 1000):
            cell = ws.cell(row = i, column = self.number).value
            if cell:
                rawList.append(cell)            
        calTime = []
        a = 3
        while a < len(rawList):
            calTime.append(rawList[a])
            a += 3
            time_tcl = 0
        for i in calTime:
            time = int(i[3:]) -int(i[:2])
            time_tcl += time
        return(time_tcl)
        
    def calTimeP(self, number):
        self.number = number
        wb = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\data_P.xlsx', data_only=True)
        ws = wb['datasheet']
        rawList = []
        for i in range(1, 1000):
            cell = ws.cell(row = i, column = self.number).value
            if cell:
                rawList.append(cell)            
        calTime = []
        a = 3
        while a < len(rawList):
            calTime.append(rawList[a])
            a += 3
            time_tcl = 0
        for i in calTime:
            time = int(i[3:]) -int(i[:2])
            time_tcl += time
        return(time_tcl)
    
    def calTimeSumB(self, *args):
        result = 0
        for i in args:
            result += self.calTimeB(i)
        print(result)

    def calTimeSumP(self, *args):
        result = 0
        for i in args:
            result += self.calTimeP(i)
        print(result)
    
    def labNumberB(self):
        wb1 = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\data_B.xlsx', data_only=True)
        ws1 = wb1['datasheet']
        wb2 = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\labList.xlsx', data_only=True)
        ws2 = wb2['rawdata']        

        user = []
        rawListA = []
        for i in range(1, 1000):
            cellA = ws1.cell(row = i, column = 2).value
            if cellA:
                rawListA.append(cellA)            
        a = 1
        while a < len(rawListA):
            user.append(rawListA[a])
            a += 3

        rawListB = []
        for i in range(1, 1000):
            cellB = ws1.cell(row = i, column = 3).value
            if cellB:
                rawListB.append(cellB)            
        b = 1
        while b < len(rawListB):
            user.append(rawListB[b])
            b += 3         

        rawListC = []
        for i in range(1, 1000):
            cellC = ws1.cell(row = i, column = 4).value
            if cellC:
                rawListC.append(cellC)            
        c = 1
        while c < len(rawListC):
            user.append(rawListC[c])
            c += 3

        rawListD = []
        for i in range(1, 1000):
            cellD = ws1.cell(row = i, column = 5).value
            if cellD:
                rawListD.append(cellD)            
        d = 1
        while d < len(rawListD):
            user.append(rawListD[b])
            d += 3
        
        nameList = []       
        for i in range(0, len(user)):
            a = user[i][-3:]
            if a not in nameList:
                nameList.append(a)   
                
        labList = {}
        for i in range(1, 1000):
            name = ws2.cell(row = i, column = 1).value
            lab = ws2.cell(row = i, column = 2).value
            labList[name] = lab
        
        labNumber = set()
        for i in nameList:
            if i in labList:
                labNumber.add(labList[i])

        print(nameList)
        print(labNumber)
        print(len(labNumber))
      
total = tclRoom()
total.calTimeSumB(2,3,4)
total.labNumberB()

