# -*- coding: utf-8 -*-
"""
Created on Sun Jul  8 15:05:56 2018

@author: GAYEON YOO
"""

import openpyxl

class tclRoom:
    
    
    def __init__(self):
       self.value = 0


    def roomTimeB(self, number):
        self.number = number
        wb = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\data_B.xlsx', data_only=True)
        ws = wb['datasheet']
        rawList = []
        for i in range(1, 500):
            cell = ws.cell(row = i, column = self.number).value
            if cell:
                rawList.append(cell)            
        calTime = []
        a = 3
        while a < len(rawList):
            calTime.append(rawList[a])
            a += 3
            time_tcl = 0
        for i in calTime:
            time = int(i[3:]) -int(i[:2])
            time_tcl += time
        return time_tcl
        
        
    def roomTimeP(self, number):
        self.number = number
        wb = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\data_P.xlsx', data_only=True)
        ws = wb['datasheet']
        rawList = []
        for i in range(1, 500):
            cell = ws.cell(row = i, column = self.number).value
            if cell:
                rawList.append(cell)            
        calTime = []
        a = 3
        while a < len(rawList):
            calTime.append(rawList[a])
            a += 3
            time_tcl = 0
        for i in calTime:
            time = int(i[3:]) -int(i[:2])
            time_tcl += time
        return time_tcl
    
    
    def sumTimeB(self, *args):
        result = 0
        for i in args:
            result += self.roomTimeB(i)
        print(result)


    def sumTimeP(self, *args):
        result = 0
        for i in args:
            result += self.roomTimeP(i)
        print(result)
    
    
    def listLabB(self, number):
        wb1 = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\data_B.xlsx', data_only=True)
        ws1 = wb1['datasheet']
        wb2 = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\labList.xlsx', data_only=True)
        ws2 = wb2['rawdata']   
        self.number = number

        user = []
        rawData = []
        for i in range(1, 1000):
            cell = ws1.cell(row = i, column = self.number).value
            if cell:
                rawData.append(cell)     

        a = 1
        while a < len(rawData):
            user.append(rawData[a])
            a += 3

        nameList = []       
        for i in range(0, len(user)):
            name = user[i][-3:]
            if name not in nameList:
                nameList.append(name)    
                
        rawLabList = {}
        for i in range(1, 500):
            name = ws2.cell(row = i, column = 1).value
            lab = ws2.cell(row = i, column = 2).value
            rawLabList[name] = lab
        
        labList = set()
        for i in nameList:
            if i in rawLabList:
                labList.add(rawLabList[i])
        print(labList)
        return list(labList)

  
    def listLabP(self, number):
        wb1 = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\data_P.xlsx', data_only=True)
        ws1 = wb1['datasheet']
        wb2 = openpyxl.load_workbook('C:\workspace\easyLife\easyTCL\labList.xlsx', data_only=True)
        ws2 = wb2['rawdata']   
        self.number = number

        user = []
        rawData = []
        for i in range(1, 1000):
            cell = ws1.cell(row = i, column = self.number).value
            if cell:
                rawData.append(cell)     

        a = 1
        while a < len(rawData):
            user.append(rawData[a])
            a += 3

        nameList = []       
        for i in range(0, len(user)):
            name = user[i][-3:]
            if name not in nameList:
                nameList.append(name)    
                
        rawLabList = {}
        for i in range(1, 500):
            name = ws2.cell(row = i, column = 1).value
            lab = ws2.cell(row = i, column = 2).value
            rawLabList[name] = lab
        
        labList = set()
        for i in nameList:
            if i in rawLabList:
                labList.add(rawLabList[i])
        print(labList)
        return list(labList)
    
    
    def sumLabB(self, *args):
        sumLab = []
        for i in args:
            sumLab += self.listLabB(i)
        sumLab2 = list(set(sumLab))    

        result = {}
        result[len(sumLab2)] = sumLab2
        print(result)
    #왜 2,3,4 넣은 값이 각각 보일까
    
    def sumLabP(self, *args):
        sumLab = []
        for i in args:
            sumLab += self.listLabP(i)
        sumLab = list(set(sumLab))    

        result = {}
        result[len(sumLab)] = sumLab
        print(result)

  
total = tclRoom()
total.sumTimeB(2,3,4)
total.sumLabB(2,3,4)

