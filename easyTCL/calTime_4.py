# -*- coding: utf-8 -*-
"""
Created on Sun Jul  8 15:05:56 2018

@author: GAYEON YOO
"""
 
import openpyxl

inputList = ['data_B.xlsx', 'data_P.xlsx']
# indexB = [2,3,4]
# indexP = [2,3,4]

class tclRoom:
    
    
    def __init__(self):
       self.value = 0


    def calTime(self, number, inputFile):
        self.number = number
        wb = openpyxl.load_workbook(inputFile, data_only=True)
        ws = wb['datasheet']
        
        rawList = []
        for i in range(1, 500):
            cell = ws.cell(row = i, column = self.number).value
            if cell:
                rawList.append(cell)            
        calTime = []
        a = 3
        while a < len(rawList):
            calTime.append(rawList[a])
            a += 3
            time_tcl = 0
        for i in calTime:
            time = int(i[3:]) -int(i[:2])
            time_tcl += time
        return time_tcl
    
    
    def sumTime(self, *args):
        result = 0
        for inputFile in inputList:
            for i in args:
                result += self.calTime(i, inputFile)
                print(inputFile + str(i) + ": " + str(self.calTime(i, inputFile)))
            print(inputFile + ": " + str(result))
            result = 0
    
    
    def listLab(self, number, inputFile):
        self.number = number
        wb1 = openpyxl.load_workbook(inputFile, data_only=True)
        ws1 = wb1['datasheet']
        wb2 = openpyxl.load_workbook('labList.xlsx', data_only=True)
        ws2 = wb2['rawdata']   

        user = []
        rawData = []
        for i in range(1, 1000):
            cell = ws1.cell(row = i, column = self.number).value
            if cell:
                rawData.append(cell)     

        a = 1
        while a < len(rawData):
            user.append(rawData[a])
            a += 3

        nameList = []       
        for i in range(0, len(user)):
            name = user[i][-3:]
            if name not in nameList:
                nameList.append(name)    
                
        rawLabList = {}
        for i in range(1, 500):
            name = ws2.cell(row = i, column = 1).value
            lab = ws2.cell(row = i, column = 2).value
            rawLabList[name] = lab
        
        labList = set()
        for i in nameList:
            if i in rawLabList:
                labList.add(rawLabList[i])
        return list(labList)

    
    def sumLab(self, *args):
        sumLab = []
        for inputFile in inputList:
            for i in args:
                sumLab += self.listLab(i, inputFile)
                sumLab2 = list(set(sumLab))    
            print(inputFile + "Lab list is: " + str(len(sumLab2)))


total = tclRoom()
total.sumTime(2,3,4)
total.sumLab(2,3,4)